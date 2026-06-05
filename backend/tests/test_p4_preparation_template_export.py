from __future__ import annotations

import os
from io import BytesIO

import httpx
from openpyxl import load_workbook


async def login_headers(client: httpx.AsyncClient) -> dict[str, str]:
    response = await client.post(
        "/api/v1/auth/login",
        json={
            "username": os.environ["BOOTSTRAP_ROOT_USERNAME"],
            "password": os.environ["BOOTSTRAP_ROOT_PASSWORD"],
        },
    )
    assert response.status_code == 200
    token = response.json()["data"]["accessToken"]
    return {"Authorization": f"Bearer {token}"}


async def _create_material_and_inventory(
    client: httpx.AsyncClient,
    headers: dict[str, str],
) -> tuple[dict, dict]:
    material_response = await client.post(
        "/api/v1/materials",
        headers=headers,
        json={
            "materialGrade": "Q355",
            "thickness": 3.0,
            "specDescription": "Preparation test sheet",
        },
    )
    assert material_response.status_code == 200
    material = material_response.json()["data"]

    inventory_response = await client.post(
        "/api/v1/inventory-items",
        headers=headers,
        json={
            "materialId": material["id"],
            "inventoryType": "leftover",
            "width": 1500,
            "length": 900,
            "thickness": 3.0,
            "quantity": 1,
            "source": "previous-leftover",
            "location": "R-01",
            "reusable": True,
        },
    )
    assert inventory_response.status_code == 200
    return material, inventory_response.json()["data"]


async def test_preparation_order_exports_template_xlsx(client: httpx.AsyncClient) -> None:
    headers = await login_headers(client)
    material, inventory = await _create_material_and_inventory(client, headers)

    order_response = await client.post(
        "/api/v1/cutting-preparations",
        headers=headers,
        json={"preparationDate": "2026-06-04"},
    )
    assert order_response.status_code == 200
    order = order_response.json()["data"]
    assert order["status"] == "draft"
    assert order["items"] == []

    item_response = await client.post(
        f"/api/v1/cutting-preparations/{order['id']}/items",
        headers=headers,
        json={
            "sourceInventoryItemId": inventory["id"],
            "sheetName": "Q355-3mm-leftover-001",
            "drawingPath": r"D:\laser\A001.dxf",
            "width": 1200,
            "length": 800,
            "materialGrade": material["materialGrade"],
            "thickness": material["thickness"],
            "quantity": 1,
            "sortOrder": 1,
        },
    )
    assert item_response.status_code == 200
    order_with_item = item_response.json()["data"]
    assert len(order_with_item["items"]) == 1

    reserved_response = await client.get(
        "/api/v1/inventory-items",
        headers=headers,
        params={"status": "reserved"},
    )
    assert reserved_response.status_code == 200
    reserved_items = reserved_response.json()["data"]
    assert [item["id"] for item in reserved_items] == [inventory["id"]]

    export_response = await client.post(
        f"/api/v1/cutting-preparations/{order['id']}/export-template",
        headers=headers,
    )
    assert export_response.status_code == 200
    export = export_response.json()["data"]
    assert export["rowCount"] == 1
    assert export["fileName"].endswith(".xlsx")
    assert export["downloadUrl"].endswith(f"/{export['id']}/download")

    generated_response = await client.get(
        f"/api/v1/cutting-preparations/{order['id']}",
        headers=headers,
    )
    assert generated_response.status_code == 200
    assert generated_response.json()["data"]["status"] == "generated"

    download_response = await client.get(export["downloadUrl"], headers=headers)
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content), data_only=True)
    sheet = workbook.active

    assert [sheet.cell(row=1, column=column).value for column in range(1, 8)] == [
        "板材名称",
        "图纸路径",
        "宽",
        "长",
        "材质",
        "厚度",
        "数量",
    ]
    assert [sheet.cell(row=2, column=column).value for column in range(1, 8)] == [
        "Q355-3mm-leftover-001",
        r"D:\laser\A001.dxf",
        1200,
        800,
        "Q355",
        3,
        1,
    ]


async def test_preparation_item_rejects_unusable_source(client: httpx.AsyncClient) -> None:
    headers = await login_headers(client)
    _, inventory = await _create_material_and_inventory(client, headers)

    order_response = await client.post(
        "/api/v1/cutting-preparations",
        headers=headers,
        json={"preparationDate": "2026-06-04"},
    )
    assert order_response.status_code == 200
    order = order_response.json()["data"]

    response = await client.post(
        f"/api/v1/cutting-preparations/{order['id']}/items",
        headers=headers,
        json={
            "sourceInventoryItemId": inventory["id"],
            "sheetName": "too-large",
            "drawingPath": "",
            "width": 1600,
            "length": 800,
            "materialGrade": "Q355",
            "thickness": 3.0,
            "quantity": 1,
        },
    )

    assert response.status_code == 200
    assert response.json()["code"] == 400
    assert response.json()["errorCode"] == "invalid_preparation_source"
