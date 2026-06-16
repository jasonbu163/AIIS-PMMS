from __future__ import annotations

import os
from io import BytesIO

import httpx
from openpyxl import Workbook, load_workbook
from sqlalchemy import Unicode

from app.cutting.models.preparation import CuttingPreparationOrder
from app.cutting.models.preparation import CuttingPreparationItem
from app.cutting.models.preparation import CuttingTemplateExport
from app.material.models.inventory_item import MaterialInventoryItem
from app.material.models.material import Material
from app.user.models.auth_token_revocation import AuthTokenRevocation
from app.user.models.user import User


async def login_headers(client: httpx.AsyncClient) -> dict[str, str]:
    response = await client.post(
        "/api/v1/auth/login",
        json={
            "username": os.environ["BOOTSTRAP_ROOT_USERNAME"],
            "password": os.environ["BOOTSTRAP_ROOT_PASSWORD"],
        },
    )
    assert response.status_code == 200
    token = response.json()["data"]["accessToken"]
    return {"Authorization": f"Bearer {token}"}


def test_chinese_business_text_columns_use_unicode() -> None:
    columns = [
        Material.__table__.c.material_grade,
        Material.__table__.c.spec_description,
        Material.__table__.c.default_unit,
        MaterialInventoryItem.__table__.c.inventory_code,
        MaterialInventoryItem.__table__.c.remark,
        MaterialInventoryItem.__table__.c.source,
        MaterialInventoryItem.__table__.c.location,
        CuttingPreparationOrder.__table__.c.created_by,
        CuttingPreparationItem.__table__.c.sheet_name,
        CuttingPreparationItem.__table__.c.drawing_path,
        CuttingPreparationItem.__table__.c.material_grade,
        CuttingTemplateExport.__table__.c.file_name,
        CuttingTemplateExport.__table__.c.file_path,
        CuttingTemplateExport.__table__.c.created_by,
        User.__table__.c.username,
        User.__table__.c.display_name,
        AuthTokenRevocation.__table__.c.reason,
    ]

    assert all(isinstance(column.type, Unicode) for column in columns)


async def test_material_and_inventory_crud(client: httpx.AsyncClient) -> None:
    headers = await login_headers(client)

    material_response = await client.post(
        "/api/v1/materials",
        headers=headers,
        json={
            "materialGrade": "Q235",
            "thickness": 2.5,
            "specDescription": "Laser cutting sheet",
            "defaultUnit": "sheet",
        },
    )
    assert material_response.status_code == 200
    material = material_response.json()["data"]
    assert material["materialGrade"] == "Q235"
    assert material["thickness"] == 2.5
    assert material["enabled"] is True

    material_page_response = await client.get(
        "/api/v1/materials/page",
        headers=headers,
        params={"page": 1, "pageSize": 20, "materialGrade": "Q"},
    )
    assert material_page_response.status_code == 200
    material_page = material_page_response.json()["data"]
    assert material_page["meta"] == {"page": 1, "pageSize": 20, "total": 1}
    assert material_page["items"][0]["id"] == material["id"]

    material_detail_response = await client.get(
        f"/api/v1/materials/{material['id']}",
        headers=headers,
    )
    assert material_detail_response.status_code == 200
    assert material_detail_response.json()["data"]["id"] == material["id"]

    material_update_response = await client.patch(
        f"/api/v1/materials/{material['id']}",
        headers=headers,
        json={
            "specDescription": "中文板材规格",
            "defaultUnit": "张",
            "enabled": False,
        },
    )
    assert material_update_response.status_code == 200
    updated_material = material_update_response.json()["data"]
    assert updated_material["specDescription"] == "中文板材规格"
    assert updated_material["defaultUnit"] == "张"
    assert updated_material["enabled"] is False

    inventory_response = await client.post(
        "/api/v1/inventory-items",
        headers=headers,
        json={
            "materialId": material["id"],
            "inventoryType": "leftover",
            "width": 1200,
            "length": 800,
            "thickness": 2.5,
            "quantity": 1,
            "source": "production-report-001",
            "location": "A-01",
            "reusable": True,
        },
    )
    assert inventory_response.status_code == 200
    inventory = inventory_response.json()["data"]
    assert inventory["inventoryType"] == "leftover"
    assert inventory["status"] == "available"
    assert inventory["materialGrade"] == "Q235"
    assert inventory["location"] == "A-01"
    assert inventory["inventoryCode"].startswith("RM:Q235-1200x800x2.5-")
    assert inventory["createdAt"]
    assert inventory["updatedAt"]

    blocked_grade_update_response = await client.patch(
        f"/api/v1/materials/{material['id']}",
        headers=headers,
        json={"materialGrade": "Q235B"},
    )
    assert blocked_grade_update_response.status_code == 200
    assert blocked_grade_update_response.json()["errorCode"] == "material_in_use"

    blocked_thickness_update_response = await client.patch(
        f"/api/v1/materials/{material['id']}",
        headers=headers,
        json={"thickness": 3.0},
    )
    assert blocked_thickness_update_response.status_code == 200
    assert blocked_thickness_update_response.json()["errorCode"] == "material_in_use"

    query_response = await client.get(
        "/api/v1/inventory-items",
        headers=headers,
        params={
            "materialId": material["id"],
            "inventoryType": "leftover",
            "status": "available",
            "minWidth": 1000,
            "minLength": 600,
            "reusable": "true",
        },
    )
    assert query_response.status_code == 200
    items = query_response.json()["data"]
    assert len(items) == 1
    assert items[0]["id"] == inventory["id"]
    assert items[0]["createdAt"]
    assert items[0]["updatedAt"]

    page_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={
            "page": 1,
            "pageSize": 20,
            "inventoryType": "leftover",
            "status": "available",
        },
    )
    assert page_response.status_code == 200
    page_data = page_response.json()["data"]
    assert page_data["meta"] == {"page": 1, "pageSize": 20, "total": 1}
    assert [item["id"] for item in page_data["items"]] == [inventory["id"]]
    assert page_data["items"][0]["createdAt"]
    assert page_data["items"][0]["updatedAt"]

    fuzzy_grade_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={
            "page": 1,
            "pageSize": 20,
            "materialGrade": "Q2",
        },
    )
    assert fuzzy_grade_response.status_code == 200
    assert [item["id"] for item in fuzzy_grade_response.json()["data"]["items"]] == [
        inventory["id"]
    ]

    fuzzy_code_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={
            "page": 1,
            "pageSize": 20,
            "inventoryCode": "1200x800",
        },
    )
    assert fuzzy_code_response.status_code == 200
    assert [item["id"] for item in fuzzy_code_response.json()["data"]["items"]] == [
        inventory["id"]
    ]

    update_response = await client.patch(
        f"/api/v1/inventory-items/{inventory['id']}",
        headers=headers,
        json={
            "quantity": 2,
            "location": "B-02",
            "status": "reserved",
        },
    )
    assert update_response.status_code == 200
    updated = update_response.json()["data"]
    assert updated["quantity"] == 2
    assert updated["location"] == "B-02"
    assert updated["status"] == "reserved"

    void_response = await client.post(
        f"/api/v1/inventory-items/{inventory['id']}/void",
        headers=headers,
    )
    assert void_response.status_code == 200
    assert void_response.json()["data"]["status"] == "voided"


async def test_inventory_depletion_and_same_spec_stock_in(
    client: httpx.AsyncClient,
) -> None:
    headers = await login_headers(client)

    material_response = await client.post(
        "/api/v1/materials",
        headers=headers,
        json={
            "materialGrade": "Q355",
            "thickness": 6,
            "specDescription": "Stock-in regression sheet",
            "defaultUnit": "sheet",
        },
    )
    assert material_response.status_code == 200
    material = material_response.json()["data"]

    inventory_response = await client.post(
        "/api/v1/inventory-items",
        headers=headers,
        json={
            "materialId": material["id"],
            "inventoryType": "whole_sheet",
            "width": 1500,
            "length": 6000,
            "thickness": 6,
            "quantity": 1,
            "source": "manual-entry",
            "location": "A-01",
        },
    )
    assert inventory_response.status_code == 200
    inventory = inventory_response.json()["data"]
    assert inventory["status"] == "available"

    depleted_response = await client.patch(
        f"/api/v1/inventory-items/{inventory['id']}",
        headers=headers,
        json={"quantity": 0},
    )
    assert depleted_response.status_code == 200
    depleted = depleted_response.json()["data"]
    assert depleted["quantity"] == 0
    assert depleted["status"] == "consumed"

    stock_in_response = await client.post(
        f"/api/v1/inventory-items/{inventory['id']}/stock-in",
        headers=headers,
        json={
            "quantity": 3,
            "source": "site-stock-in",
            "location": "A-02",
        },
    )
    assert stock_in_response.status_code == 200
    stocked = stock_in_response.json()["data"]
    assert stocked["id"] == inventory["id"]
    assert stocked["quantity"] == 3
    assert stocked["status"] == "available"
    assert stocked["source"] == "site-stock-in"
    assert stocked["location"] == "A-02"

    reserved_response = await client.patch(
        f"/api/v1/inventory-items/{inventory['id']}",
        headers=headers,
        json={"status": "reserved"},
    )
    assert reserved_response.status_code == 200
    assert reserved_response.json()["data"]["status"] == "reserved"

    blocked_stock_in_response = await client.post(
        f"/api/v1/inventory-items/{inventory['id']}/stock-in",
        headers=headers,
        json={"quantity": 1},
    )
    assert blocked_stock_in_response.status_code == 200
    assert blocked_stock_in_response.json()["errorCode"] == "invalid_inventory_status"

    consume_response = await client.post(
        f"/api/v1/inventory-items/{inventory['id']}/consume",
        headers=headers,
        json={
            "quantity": 2,
            "source": "site-consume",
            "remark": "现场领用",
        },
    )
    assert consume_response.status_code == 200
    consumed = consume_response.json()["data"]
    assert consumed["quantity"] == 1
    assert consumed["status"] == "reserved"
    assert consumed["source"] == "site-consume"
    assert consumed["remark"] == "现场领用"

    consume_to_zero_response = await client.post(
        f"/api/v1/inventory-items/{inventory['id']}/consume",
        headers=headers,
        json={"quantity": 1},
    )
    assert consume_to_zero_response.status_code == 200
    consumed_to_zero = consume_to_zero_response.json()["data"]
    assert consumed_to_zero["quantity"] == 0
    assert consumed_to_zero["status"] == "consumed"

    blocked_over_consume_response = await client.post(
        f"/api/v1/inventory-items/{inventory['id']}/consume",
        headers=headers,
        json={"quantity": 1},
    )
    assert blocked_over_consume_response.status_code == 200
    assert blocked_over_consume_response.json()["errorCode"] == "invalid_inventory_quantity"

    page_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={"materialId": material["id"], "page": 1, "pageSize": 20},
    )
    assert page_response.status_code == 200
    page_data = page_response.json()["data"]
    assert page_data["meta"]["total"] == 1
    assert page_data["items"][0]["id"] == inventory["id"]


def build_inventory_xlsx(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "板材名称",
            "图纸路径",
            "宽",
            "长",
            "材质",
            "厚度",
            "数量",
            "使用数量",
            "新增数量",
        ]
    )
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def test_inventory_xlsx_import_export_and_code_lookup(
    client: httpx.AsyncClient,
) -> None:
    headers = await login_headers(client)
    content = build_inventory_xlsx(
        [
            ["Sheet-A", "", 1200, 800, "Q235", 3, 10, "", ""],
            ["Sheet-B", "/drawings/sheet-b.dxf", 1500, 600, "Q345", 4, 2, "", ""],
        ]
    )

    dry_run_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "true"},
        files={
            "file": (
                "inventory.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert dry_run_response.status_code == 200
    dry_run = dry_run_response.json()["data"]
    assert dry_run["dryRun"] is True
    assert dry_run["created"] == 2
    assert dry_run["updated"] == 0
    assert dry_run["errors"] == []
    assert dry_run["previewRows"][0]["usedQuantity"] == 0
    assert dry_run["previewRows"][0]["addedQuantity"] == 0
    assert dry_run["previewRows"][0]["action"] == "create"
    assert "库存数 10" in dry_run["previewRows"][0]["remark"]

    import_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "false"},
        files={
            "file": (
                "inventory.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    imported = import_response.json()["data"]
    assert imported["created"] == 2
    assert imported["updated"] == 0

    page_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={"page": 1, "pageSize": 20},
    )
    assert page_response.status_code == 200
    items = page_response.json()["data"]["items"]
    q235_item = next(item for item in items if item["materialGrade"] == "Q235")
    assert q235_item["inventoryCode"].startswith("RM:Q235-1200x800x3-")
    assert q235_item["quantity"] == 10
    assert "未匹配到库存" in q235_item["remark"]

    lookup_response = await client.get(
        "/api/v1/inventory-items/by-code",
        headers=headers,
        params={"inventoryCode": q235_item["inventoryCode"]},
    )
    assert lookup_response.status_code == 200
    assert lookup_response.json()["data"]["id"] == q235_item["id"]
    assert lookup_response.json()["data"]["createdAt"]
    assert lookup_response.json()["data"]["updatedAt"]

    update_response = await client.patch(
        f"/api/v1/inventory-items/{q235_item['id']}",
        headers=headers,
        json={
            "quantity": 8,
            "remark": "manual review",
            "location": "C-09",
            "status": "reserved",
        },
    )
    assert update_response.status_code == 200
    update_result = update_response.json()["data"]
    assert update_result["remark"] == "manual review"

    updated_lookup_response = await client.get(
        "/api/v1/inventory-items/by-code",
        headers=headers,
        params={"inventoryCode": q235_item["inventoryCode"]},
    )
    assert updated_lookup_response.status_code == 200
    updated_item = updated_lookup_response.json()["data"]
    assert updated_item["quantity"] == 8
    assert updated_item["location"] == "C-09"
    assert updated_item["status"] == "reserved"

    consume_content = build_inventory_xlsx(
        [
            ["Ignored-A", "/ignored/a.dxf", 1200, 800, "Q235", 3, 99, 3, ""],
            ["Ignored-B", "/ignored/b.dxf", 1500, 600, "Q345", 4, 99, 5, ""],
        ]
    )
    consume_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "false"},
        files={
            "file": (
                "inventory-consume.xlsx",
                consume_content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert consume_response.status_code == 200
    consumed = consume_response.json()["data"]
    assert consumed["created"] == 0
    assert consumed["updated"] == 2

    consumed_lookup_response = await client.get(
        "/api/v1/inventory-items/by-code",
        headers=headers,
        params={"inventoryCode": q235_item["inventoryCode"]},
    )
    assert consumed_lookup_response.status_code == 200
    consumed_item = consumed_lookup_response.json()["data"]
    assert consumed_item["quantity"] == 5
    assert consumed_item["remark"] == "本次操作使用数量 3，库存数 8 - 3 = 5。"

    stock_in_content = build_inventory_xlsx(
        [
            ["Ignored-A", "/ignored/a.dxf", 1200, 800, "Q235", 3, 99, "", 4],
        ]
    )
    stock_in_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "false"},
        files={
            "file": (
                "inventory-stock-in.xlsx",
                stock_in_content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert stock_in_response.status_code == 200
    assert stock_in_response.json()["data"]["updated"] == 1

    stocked_lookup_response = await client.get(
        "/api/v1/inventory-items/by-code",
        headers=headers,
        params={"inventoryCode": q235_item["inventoryCode"]},
    )
    assert stocked_lookup_response.status_code == 200
    stocked_item = stocked_lookup_response.json()["data"]
    assert stocked_item["quantity"] == 9
    assert stocked_item["status"] == "available"
    assert stocked_item["remark"] == "本次入库数量 4，库存数 5 + 4 = 9。"

    export_response = await client.post(
        "/api/v1/inventory-items/export-xlsx",
        headers=headers,
        json={"inventoryCodes": [q235_item["inventoryCode"]]},
    )
    assert export_response.status_code == 200
    workbook = load_workbook(BytesIO(export_response.content), read_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "板材名称",
        "图纸路径",
        "宽",
        "长",
        "材质",
        "厚度",
        "数量",
        "使用数量",
        "新增数量",
    )
    assert rows[1][0] == q235_item["inventoryCode"]
    assert rows[1][1] in (None, "")
    assert rows[1][4] == "Q235"
    assert rows[1][6] == 9
    assert rows[1][7] in (None, "")
    assert rows[1][8] in (None, "")


async def test_inventory_xlsx_consumes_to_zero_and_restock_reactivates(
    client: httpx.AsyncClient,
) -> None:
    headers = await login_headers(client)
    initial_content = build_inventory_xlsx(
        [
            ["Sheet-Z", "", 900, 500, "Q460", 8, 1, 0],
        ]
    )

    initial_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "false"},
        files={
            "file": (
                "inventory-initial.xlsx",
                initial_content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert initial_response.status_code == 200
    assert initial_response.json()["data"]["created"] == 1

    consume_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "false"},
        files={
            "file": (
                "inventory-consume-zero.xlsx",
                build_inventory_xlsx([["Sheet-Z", "", 900, 500, "Q460", 8, "", 1, ""]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert consume_response.status_code == 200
    assert consume_response.json()["data"]["updated"] == 1

    consumed_page_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={"materialGrade": "Q460", "thickness": 8, "page": 1, "pageSize": 20},
    )
    assert consumed_page_response.status_code == 200
    consumed_item = consumed_page_response.json()["data"]["items"][0]
    assert consumed_item["quantity"] == 0
    assert consumed_item["status"] == "consumed"

    stock_in_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "false"},
        files={
            "file": (
                "inventory-stock-in.xlsx",
                build_inventory_xlsx([["Sheet-Z", "", 900, 500, "Q460", 8, "", "", 2]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert stock_in_response.status_code == 200
    assert stock_in_response.json()["data"]["created"] == 0
    assert stock_in_response.json()["data"]["updated"] == 1

    stocked_page_response = await client.get(
        "/api/v1/inventory-items/page",
        headers=headers,
        params={"materialGrade": "Q460", "thickness": 8, "page": 1, "pageSize": 20},
    )
    assert stocked_page_response.status_code == 200
    page_data = stocked_page_response.json()["data"]
    assert page_data["meta"]["total"] == 1
    stocked_item = page_data["items"][0]
    assert stocked_item["id"] == consumed_item["id"]
    assert stocked_item["quantity"] == 2
    assert stocked_item["status"] == "available"


async def test_inventory_xlsx_import_action_validation(
    client: httpx.AsyncClient,
) -> None:
    headers = await login_headers(client)
    material_response = await client.post(
        "/api/v1/materials",
        headers=headers,
        json={
            "materialGrade": "Q500",
            "thickness": 10,
            "specDescription": "Validation sheet",
            "defaultUnit": "sheet",
        },
    )
    assert material_response.status_code == 200
    material = material_response.json()["data"]
    existing_response = await client.post(
        "/api/v1/inventory-items",
        headers=headers,
        json={
            "materialId": material["id"],
            "inventoryType": "leftover",
            "width": 1000,
            "length": 500,
            "thickness": 10,
            "quantity": 5,
        },
    )
    assert existing_response.status_code == 200

    invalid_content = build_inventory_xlsx(
        [
            ["Existing-Quantity-Only", "", 1000, 500, "Q500", 10, 2, "", ""],
            ["Missing-Consume", "", 2000, 500, "Q501", 10, "", 1, ""],
            ["Missing-Stock-In", "", 2000, 600, "Q502", 10, "", "", 1],
            ["Ambiguous-Use-And-Add", "", 1000, 500, "Q500", 10, "", 1, 1],
            ["Ambiguous-All", "", 1000, 500, "Q500", 10, 3, 1, 1],
            ["Empty-Action", "", 2000, 700, "Q503", 10, "", "", ""],
        ]
    )
    invalid_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "true"},
        files={
            "file": (
                "inventory-invalid-actions.xlsx",
                invalid_content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert invalid_response.status_code == 200
    invalid = invalid_response.json()["data"]
    assert invalid["validRows"] == 0
    assert invalid["skipped"] == 6
    messages = {error["rowNumber"]: error["message"] for error in invalid["errors"]}
    assert "规格已存在" in messages[2]
    assert "未匹配到库存" in messages[3]
    assert "未匹配到库存" in messages[4]
    assert "不能同时填写使用数量和新增数量" in messages[5]
    assert "不能同时填写使用数量和新增数量" in messages[6]
    assert "数量、使用数量、新增数量至少填写一项" in messages[7]


async def test_inventory_xlsx_limits(client: httpx.AsyncClient) -> None:
    headers = await login_headers(client)
    too_many_rows = [
        ["Sheet-A", "/drawings/sheet-a.dxf", 1200, 800, "Q235", 3, 1, 0]
        for _ in range(201)
    ]
    import_response = await client.post(
        "/api/v1/inventory-items/import-xlsx",
        headers=headers,
        params={"dryRun": "true"},
        files={
            "file": (
                "inventory-too-large.xlsx",
                build_inventory_xlsx(too_many_rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["errorCode"] == "inventory_xlsx_limit_exceeded"

    export_response = await client.post(
        "/api/v1/inventory-items/export-xlsx",
        headers=headers,
        json={"inventoryCodes": [f"RM:TEST-{index}" for index in range(201)]},
    )
    assert export_response.status_code == 200
    assert export_response.json()["errorCode"] == "inventory_xlsx_limit_exceeded"


async def test_material_inventory_requires_auth(client: httpx.AsyncClient) -> None:
    response = await client.get("/api/v1/inventory-items")

    assert response.status_code == 401
