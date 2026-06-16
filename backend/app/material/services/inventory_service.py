from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.material.cruds import inventory_crud, material_crud
from app.material.models.inventory_item import MaterialInventoryItem
from app.material.models.material import Material
from app.material.schemas.inventory import (
    InventoryItemCreateIn,
    InventoryItemConsumeIn,
    InventoryItemOut,
    InventoryItemStockInIn,
    InventoryItemUpdateIn,
)
from app.material.schemas.inventory import (
    InventoryImportError,
    InventoryImportPreviewRow,
    InventoryImportResult,
)
from common.error_codes import ErrorCode
from common.exceptions import BusinessException
from common.log import logger
from common.pagination import PageData, PageMeta, normalize_page

ALLOWED_STATUS = {"available", "reserved", "consumed", "scrapped", "voided"}
QUANTITY_MANAGED_STATUS = {"available", "reserved", "consumed"}
RESTOCK_EXCLUDED_STATUS = {"reserved", "scrapped", "voided"}
CONSUME_EXCLUDED_STATUS = {"scrapped", "voided"}
STOCK_IN_ALLOWED_STATUS = {"available", "consumed"}
CONSUME_ALLOWED_STATUS = {"available", "reserved"}
INVENTORY_XLSX_ROW_LIMIT = 200
INVENTORY_EXPORT_HEADERS = [
    ("inventory_code", "板材名称"),
    ("empty_drawing_path", "图纸路径"),
    ("width", "宽"),
    ("length", "长"),
    ("material_grade", "材质"),
    ("thickness", "厚度"),
    ("quantity", "数量"),
    ("empty_used_quantity", "使用数量"),
    ("empty_added_quantity", "新增数量"),
]
HEADER_ALIASES = {
    "materialgrade": "material_grade",
    "material_grade": "material_grade",
    "材质": "material_grade",
    "width": "width",
    "宽": "width",
    "length": "length",
    "长": "length",
    "thickness": "thickness",
    "厚度": "thickness",
    "quantity": "quantity",
    "数量": "quantity",
    "usedquantity": "used_quantity",
    "used_quantity": "used_quantity",
    "usequantity": "used_quantity",
    "使用数量": "used_quantity",
    "addedquantity": "added_quantity",
    "added_quantity": "added_quantity",
    "addquantity": "added_quantity",
    "新增数量": "added_quantity",
}


def to_inventory_out(item: MaterialInventoryItem, material_grade: str) -> InventoryItemOut:
    return InventoryItemOut(
        id=item.id,
        inventory_code=item.inventory_code,
        material_id=item.material_id,
        material_grade=material_grade,
        inventory_type=item.inventory_type,
        width=item.width,
        length=item.length,
        thickness=item.thickness,
        quantity=item.quantity,
        remark=item.remark,
        source=item.source,
        location=item.location,
        status=item.status,
        reusable=item.reusable,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _format_dimension(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return str(value).rstrip("0").rstrip(".")


def _sanitize_code_part(value: str) -> str:
    cleaned = value.strip().replace(" ", "")
    for char in (":", "/", "\\"):
        cleaned = cleaned.replace(char, "-")
    return cleaned or "UNKNOWN"


def build_inventory_code(
    *,
    material_grade: str,
    width: float,
    length: float,
    thickness: float,
    inventory_item_id: int,
) -> str:
    date_key = datetime.now().strftime("%Y%m%d")
    grade = _sanitize_code_part(material_grade)
    size = "x".join(
        [
            _format_dimension(width),
            _format_dimension(length),
            _format_dimension(thickness),
        ]
    )
    return f"RM:{grade}-{size}-{date_key}-{inventory_item_id}"


def _temporary_inventory_code() -> str:
    return f"RM:PENDING-{uuid4().hex[:16]}"


async def _ensure_material(
    db: AsyncSession,
    *,
    material_grade: str,
    thickness: float,
) -> Material:
    material = await material_crud.get_material_by_grade_thickness(
        db,
        material_grade,
        thickness,
    )
    if material is not None:
        return material
    return await material_crud.create_material(
        db,
        Material(
            material_grade=material_grade,
            thickness=thickness,
            spec_description="",
            default_unit="sheet",
            enabled=True,
        ),
    )


async def _assign_generated_inventory_code(
    item: MaterialInventoryItem,
    *,
    material_grade: str,
) -> None:
    item.inventory_code = build_inventory_code(
        material_grade=material_grade,
        width=item.width,
        length=item.length,
        thickness=item.thickness,
        inventory_item_id=item.id,
    )


def _normalize_status_for_quantity(quantity: int, status: str) -> str:
    if quantity <= 0 and status in QUANTITY_MANAGED_STATUS:
        return "consumed"
    if quantity > 0 and status == "consumed":
        return "available"
    return status


async def create_inventory_item(
    db: AsyncSession,
    data: InventoryItemCreateIn,
) -> InventoryItemOut:
    material = await material_crud.get_material(db, data.material_id)
    if material is None:
        raise BusinessException(ErrorCode.MATERIAL_NOT_FOUND)

    inventory_code = data.inventory_code.strip() if data.inventory_code else ""
    if inventory_code:
        existing = await inventory_crud.get_inventory_item_by_code(db, inventory_code)
        if existing is not None:
            raise BusinessException(ErrorCode.INVENTORY_CODE_ALREADY_EXISTS)

    if not inventory_code:
        existing = await inventory_crud.get_inventory_item_by_spec(
            db,
            material_id=data.material_id,
            inventory_type=data.inventory_type,
            width=data.width,
            length=data.length,
            thickness=data.thickness,
            excluded_statuses=RESTOCK_EXCLUDED_STATUS,
        )
        if existing is not None:
            old_status = existing.status
            old_quantity = existing.quantity
            existing.quantity = max(0, existing.quantity) + data.quantity
            existing.status = _normalize_status_for_quantity(existing.quantity, "available")
            existing.remark = data.remark
            existing.source = data.source
            existing.location = data.location
            existing.reusable = data.reusable
            await db.commit()
            await db.refresh(existing)
            logger.info(
                "inventory_item_stocked_in id={} code={} old_status={} new_status={} "
                "old_quantity={} stock_in_quantity={} new_quantity={}",
                existing.id,
                existing.inventory_code,
                old_status,
                existing.status,
                old_quantity,
                data.quantity,
                existing.quantity,
            )
            return to_inventory_out(existing, material.material_grade)

    create_data = data.model_dump(exclude_none=True)
    create_data["inventory_code"] = inventory_code or _temporary_inventory_code()
    create_data["status"] = _normalize_status_for_quantity(
        int(create_data.get("quantity", 0)),
        str(create_data.get("status", "available")),
    )
    item = await inventory_crud.create_inventory_item(
        db,
        MaterialInventoryItem(**create_data),
    )
    if not inventory_code:
        await _assign_generated_inventory_code(item, material_grade=material.material_grade)
    await db.commit()
    await db.refresh(item)
    logger.info(
        "inventory_item_created id={} code={} material_id={} type={} quantity={} status={}",
        item.id,
        item.inventory_code,
        item.material_id,
        item.inventory_type,
        item.quantity,
        item.status,
    )
    return to_inventory_out(item, material.material_grade)


async def list_inventory_items(
    db: AsyncSession,
    *,
    material_id: int | None = None,
    inventory_type: str | None = None,
    status: str | None = None,
    reusable: bool | None = None,
    min_width: float | None = None,
    min_length: float | None = None,
    material_grade: str | None = None,
    thickness: float | None = None,
    inventory_code: str | None = None,
) -> list[InventoryItemOut]:
    rows = await inventory_crud.list_inventory_items(
        db,
        inventory_code=inventory_code,
        material_id=material_id,
        inventory_type=inventory_type,
        status=status,
        reusable=reusable,
        min_width=min_width,
        min_length=min_length,
        material_grade=material_grade,
        thickness=thickness,
    )
    return [to_inventory_out(item, grade) for item, grade in rows]


async def page_inventory_items(
    db: AsyncSession,
    *,
    page: int = 1,
    page_size: int = 20,
    inventory_code: str | None = None,
    material_id: int | None = None,
    inventory_type: str | None = None,
    status: str | None = None,
    reusable: bool | None = None,
    min_width: float | None = None,
    min_length: float | None = None,
    material_grade: str | None = None,
    thickness: float | None = None,
) -> PageData[InventoryItemOut]:
    page, page_size = normalize_page(page, page_size)
    total = await inventory_crud.count_inventory_items(
        db,
        inventory_code=inventory_code,
        material_id=material_id,
        inventory_type=inventory_type,
        status=status,
        reusable=reusable,
        min_width=min_width,
        min_length=min_length,
        material_grade=material_grade,
        thickness=thickness,
    )
    rows = await inventory_crud.page_inventory_items(
        db,
        page=page,
        page_size=page_size,
        inventory_code=inventory_code,
        material_id=material_id,
        inventory_type=inventory_type,
        status=status,
        reusable=reusable,
        min_width=min_width,
        min_length=min_length,
        material_grade=material_grade,
        thickness=thickness,
    )
    return PageData(
        items=[to_inventory_out(item, grade) for item, grade in rows],
        meta=PageMeta(page=page, page_size=page_size, total=total),
    )


async def get_inventory_item_by_code(db: AsyncSession, inventory_code: str) -> InventoryItemOut:
    item = await inventory_crud.get_inventory_item_by_code(db, inventory_code)
    if item is None:
        raise BusinessException(ErrorCode.INVENTORY_ITEM_NOT_FOUND)
    material = await material_crud.get_material(db, item.material_id)
    if material is None:
        raise BusinessException(ErrorCode.MATERIAL_NOT_FOUND)
    return to_inventory_out(item, material.material_grade)


async def update_inventory_item(
    db: AsyncSession,
    inventory_item_id: int,
    data: InventoryItemUpdateIn,
) -> InventoryItemOut:
    item = await inventory_crud.get_inventory_item(db, inventory_item_id)
    if item is None:
        raise BusinessException(ErrorCode.INVENTORY_ITEM_NOT_FOUND)

    old_status = item.status
    old_quantity = item.quantity
    update_data = data.model_dump(exclude_unset=True)
    status = update_data.get("status")
    if status is not None and status not in ALLOWED_STATUS:
        raise BusinessException(ErrorCode.INVALID_INVENTORY_STATUS)
    for field, value in update_data.items():
        setattr(item, field, value)
    item.status = _normalize_status_for_quantity(item.quantity, item.status)

    material = await material_crud.get_material(db, item.material_id)
    if material is None:
        raise BusinessException(ErrorCode.MATERIAL_NOT_FOUND)
    await db.commit()
    await db.refresh(item)
    logger.info(
        "inventory_item_updated id={} code={} old_status={} new_status={} "
        "old_quantity={} new_quantity={}",
        item.id,
        item.inventory_code,
        old_status,
        item.status,
        old_quantity,
        item.quantity,
    )
    return to_inventory_out(item, material.material_grade)


async def stock_in_inventory_item(
    db: AsyncSession,
    inventory_item_id: int,
    data: InventoryItemStockInIn,
) -> InventoryItemOut:
    item = await inventory_crud.get_inventory_item(db, inventory_item_id)
    if item is None:
        raise BusinessException(ErrorCode.INVENTORY_ITEM_NOT_FOUND)
    if item.status not in STOCK_IN_ALLOWED_STATUS:
        raise BusinessException(ErrorCode.INVALID_INVENTORY_STATUS)

    material = await material_crud.get_material(db, item.material_id)
    if material is None:
        raise BusinessException(ErrorCode.MATERIAL_NOT_FOUND)

    old_status = item.status
    old_quantity = item.quantity
    item.quantity = max(0, item.quantity) + data.quantity
    item.status = "available"
    item.source = data.source
    if data.location is not None:
        item.location = data.location
    if data.remark is not None:
        item.remark = data.remark
    if data.reusable is not None:
        item.reusable = data.reusable

    await db.commit()
    await db.refresh(item)
    logger.info(
        "inventory_item_stocked_in id={} code={} old_status={} new_status={} "
        "old_quantity={} stock_in_quantity={} new_quantity={}",
        item.id,
        item.inventory_code,
        old_status,
        item.status,
        old_quantity,
        data.quantity,
        item.quantity,
    )
    return to_inventory_out(item, material.material_grade)


async def consume_inventory_item(
    db: AsyncSession,
    inventory_item_id: int,
    data: InventoryItemConsumeIn,
) -> InventoryItemOut:
    item = await inventory_crud.get_inventory_item(db, inventory_item_id)
    if item is None:
        raise BusinessException(ErrorCode.INVENTORY_ITEM_NOT_FOUND)
    if data.quantity > item.quantity:
        raise BusinessException(ErrorCode.INVALID_INVENTORY_QUANTITY)
    if item.status not in CONSUME_ALLOWED_STATUS:
        raise BusinessException(ErrorCode.INVALID_INVENTORY_STATUS)

    material = await material_crud.get_material(db, item.material_id)
    if material is None:
        raise BusinessException(ErrorCode.MATERIAL_NOT_FOUND)

    old_status = item.status
    old_quantity = item.quantity
    item.quantity -= data.quantity
    item.status = _normalize_status_for_quantity(item.quantity, item.status)
    item.source = data.source
    if data.remark is not None:
        item.remark = data.remark

    await db.commit()
    await db.refresh(item)
    logger.info(
        "inventory_item_consumed id={} code={} old_status={} new_status={} "
        "old_quantity={} consume_quantity={} new_quantity={}",
        item.id,
        item.inventory_code,
        old_status,
        item.status,
        old_quantity,
        data.quantity,
        item.quantity,
    )
    return to_inventory_out(item, material.material_grade)


async def void_inventory_item(db: AsyncSession, inventory_item_id: int) -> InventoryItemOut:
    return await update_inventory_item(
        db,
        inventory_item_id,
        InventoryItemUpdateIn(status="voided"),
    )


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip()
    return text.replace(" ", "").replace("-", "").lower()


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_float(value: Any, field_name: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be a number") from exc


def _as_optional_positive_int(value: Any, field_name: str) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be an integer") from exc
    if number <= 0:
        raise ValueError(f"{field_name} must be greater than 0")
    return number


def _as_non_negative_int(value: Any, field_name: str, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be an integer") from exc
    if number < 0:
        raise ValueError(f"{field_name} must be greater than or equal to 0")
    return number


def _parse_inventory_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[InventoryImportError]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise BusinessException(ErrorCode.INVENTORY_XLSX_INVALID) from exc
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_map: dict[int, str] = {}
    for index, header in enumerate(rows[0]):
        normalized = _normalize_header(header)
        field_name = HEADER_ALIASES.get(normalized)
        if field_name is not None:
            header_map[index] = field_name

    parsed_rows: list[dict[str, Any]] = []
    errors: list[InventoryImportError] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if all(value in (None, "") for value in row):
            continue
        raw = {field: row[index] for index, field in header_map.items() if index < len(row)}
        try:
            material_grade = _as_str(raw.get("material_grade"))
            if not material_grade:
                raise ValueError("material_grade is required")
            width = _as_float(raw.get("width"), "width")
            length = _as_float(raw.get("length"), "length")
            thickness = _as_float(raw.get("thickness"), "thickness")
            parsed_rows.append(
                {
                    "row_number": row_number,
                    "inventory_code": "",
                    "material_grade": material_grade,
                    "inventory_type": "leftover",
                    "width": width,
                    "length": length,
                    "thickness": thickness,
                    "quantity": _as_optional_positive_int(raw.get("quantity"), "quantity"),
                    "used_quantity": _as_non_negative_int(
                        raw.get("used_quantity"),
                        "used_quantity",
                    ),
                    "added_quantity": _as_non_negative_int(
                        raw.get("added_quantity"),
                        "added_quantity",
                    ),
                    "source": "",
                    "location": "",
                    "status": "available",
                    "reusable": True,
                    "remark": "",
                }
            )
        except ValueError as exc:
            errors.append(InventoryImportError(row_number=row_number, message=str(exc)))
    return parsed_rows, errors


def _to_preview_row(row: dict[str, Any], action: str) -> InventoryImportPreviewRow:
    return InventoryImportPreviewRow(
        row_number=row["row_number"],
        action=action,  # type: ignore[arg-type]
        inventory_code=row["inventory_code"] or None,
        material_grade=row["material_grade"],
        inventory_type=row["inventory_type"],
        width=row["width"],
        length=row["length"],
        thickness=row["thickness"],
        quantity=row["quantity"] or 0,
        used_quantity=row["used_quantity"],
        added_quantity=row["added_quantity"],
        remark=row["remark"],
        source=row["source"],
        location=row["location"],
        status=row["status"],
        reusable=row["reusable"],
    )


def _resolve_import_row(
    row: dict[str, Any],
    *,
    matched: bool,
    current_quantity: int,
    current_status: str,
) -> tuple[str, int, str, str]:
    quantity = row["quantity"]
    used_quantity = row["used_quantity"]
    added_quantity = row["added_quantity"]

    if used_quantity > 0 and added_quantity > 0:
        raise ValueError("不能同时填写使用数量和新增数量。")
    if quantity is None and used_quantity == 0 and added_quantity == 0:
        raise ValueError("数量、使用数量、新增数量至少填写一项。")

    if used_quantity > 0:
        if not matched:
            raise ValueError("未匹配到库存，不能扣减库存。")
        if current_status not in CONSUME_ALLOWED_STATUS:
            raise ValueError("库存状态不允许扣减。")
        if used_quantity > current_quantity:
            shortage = used_quantity - current_quantity
            return (
                "consume",
                0,
                "consumed",
                f"本次操作使用数量 {used_quantity} 大于库存数 {current_quantity}，已将库存数置为0，差额{shortage}请人工确认。",
            )
        next_quantity = current_quantity - used_quantity
        return (
            "consume",
            next_quantity,
            _normalize_status_for_quantity(next_quantity, "available"),
            f"本次操作使用数量 {used_quantity}，库存数 {current_quantity} - {used_quantity} = {next_quantity}。",
        )

    if added_quantity > 0:
        if not matched:
            raise ValueError("未匹配到库存，不能新增库存。")
        if current_status not in STOCK_IN_ALLOWED_STATUS:
            raise ValueError("库存状态不允许入库。")
        next_quantity = max(0, current_quantity) + added_quantity
        return (
            "stock_in",
            next_quantity,
            "available",
            f"本次入库数量 {added_quantity}，库存数 {current_quantity} + {added_quantity} = {next_quantity}。",
        )

    if matched:
        raise ValueError("规格已存在，请填写使用数量或新增数量。")

    next_quantity = int(quantity)
    return (
        "create",
        next_quantity,
        _normalize_status_for_quantity(next_quantity, row["status"]),
        f"未匹配到库存，已新建。库存数 {next_quantity}。",
    )


async def import_inventory_xlsx(
    db: AsyncSession,
    *,
    content: bytes,
    dry_run: bool,
) -> InventoryImportResult:
    rows, errors = _parse_inventory_xlsx(content)
    total_rows = len(rows) + len(errors)
    if total_rows > INVENTORY_XLSX_ROW_LIMIT:
        raise BusinessException(ErrorCode.INVENTORY_XLSX_LIMIT_EXCEEDED)

    preview_rows: list[InventoryImportPreviewRow] = []
    resolved_rows: list[dict[str, Any]] = []
    simulated_by_spec: dict[tuple[str, str, float, float, float], dict[str, Any]] = {}
    for row in rows:
        spec_key = (
            row["material_grade"],
            row["inventory_type"],
            row["thickness"],
            row["width"],
            row["length"],
        )
        simulated = simulated_by_spec.get(spec_key)
        if simulated is None:
            material = await material_crud.get_material_by_grade_thickness(
                db,
                row["material_grade"],
                row["thickness"],
            )
            existing = None
            if material is not None:
                existing = await inventory_crud.get_inventory_item_by_spec(
                    db,
                    material_id=material.id,
                    inventory_type=row["inventory_type"],
                    width=row["width"],
                    length=row["length"],
                    thickness=row["thickness"],
                )
            matched = existing is not None
            current_quantity = existing.quantity if existing is not None else row["quantity"]
            current_status = existing.status if existing is not None else row["status"]
            row["inventory_code"] = existing.inventory_code if existing is not None else ""
        else:
            matched = True
            current_quantity = simulated["quantity"]
            current_status = simulated["status"]
            row["inventory_code"] = simulated["inventory_code"]

        try:
            action, next_quantity, next_status, remark = _resolve_import_row(
                row,
                matched=matched,
                current_quantity=current_quantity,
                current_status=current_status,
            )
        except ValueError as exc:
            errors.append(
                InventoryImportError(row_number=row["row_number"], message=str(exc))
            )
            continue
        row["remark"] = remark
        row["status"] = next_status
        row["action"] = action
        row["next_quantity"] = next_quantity
        simulated_by_spec[spec_key] = {
            "quantity": next_quantity,
            "status": next_status,
            "inventory_code": row["inventory_code"],
        }
        preview_rows.append(_to_preview_row(row, action))
        resolved_rows.append(row)

    result = InventoryImportResult(
        dry_run=dry_run,
        total_rows=total_rows,
        valid_rows=len(resolved_rows),
        created=sum(1 for row in resolved_rows if row["action"] == "create"),
        updated=sum(1 for row in resolved_rows if row["action"] in {"consume", "stock_in"}),
        skipped=len(errors),
        errors=errors,
        preview_rows=preview_rows,
    )
    if dry_run or errors:
        logger.info(
            "inventory_xlsx_import_preview dry_run={} total_rows={} valid_rows={} "
            "created={} updated={} skipped={}",
            dry_run,
            result.total_rows,
            result.valid_rows,
            result.created,
            result.updated,
            result.skipped,
        )
        return result

    created = 0
    updated = 0
    for row in resolved_rows:
        material = await _ensure_material(
            db,
            material_grade=row["material_grade"],
            thickness=row["thickness"],
        )
        if row["action"] == "create":
            item = await inventory_crud.create_inventory_item(
                db,
                MaterialInventoryItem(
                    inventory_code=row["inventory_code"] or _temporary_inventory_code(),
                    material_id=material.id,
                    inventory_type=row["inventory_type"],
                    width=row["width"],
                    length=row["length"],
                    thickness=row["thickness"],
                    quantity=row["next_quantity"],
                    remark=row["remark"],
                    source=row["source"],
                    location=row["location"],
                    status=row["status"],
                    reusable=row["reusable"],
                ),
            )
            if not row["inventory_code"]:
                await _assign_generated_inventory_code(item, material_grade=material.material_grade)
            created += 1
        else:
            existing = await inventory_crud.get_inventory_item_by_spec(
                db,
                material_id=material.id,
                inventory_type=row["inventory_type"],
                width=row["width"],
                length=row["length"],
                thickness=row["thickness"],
            )
            if existing is None:
                raise BusinessException(ErrorCode.INVENTORY_ITEM_NOT_FOUND)
            if row["action"] == "consume" and existing.status not in CONSUME_ALLOWED_STATUS:
                raise BusinessException(ErrorCode.INVALID_INVENTORY_STATUS)
            if row["action"] == "stock_in" and existing.status not in STOCK_IN_ALLOWED_STATUS:
                raise BusinessException(ErrorCode.INVALID_INVENTORY_STATUS)
            existing.quantity = row["next_quantity"]
            existing.status = row["status"]
            existing.remark = row["remark"]
            if row["action"] == "stock_in":
                existing.source = row["source"]
                existing.location = row["location"]
                existing.reusable = row["reusable"]
            updated += 1

    await db.commit()
    result.created = created
    result.updated = updated
    logger.info(
        "inventory_xlsx_import_applied total_rows={} valid_rows={} created={} updated={} skipped={}",
        result.total_rows,
        result.valid_rows,
        result.created,
        result.updated,
        result.skipped,
    )
    return result


async def export_inventory_xlsx(
    db: AsyncSession,
    *,
    inventory_codes: list[str],
) -> bytes:
    codes = []
    seen: set[str] = set()
    for code in inventory_codes:
        cleaned = code.strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            codes.append(cleaned)
    if not codes:
        raise BusinessException(ErrorCode.INVENTORY_EXPORT_EMPTY)
    if len(codes) > INVENTORY_XLSX_ROW_LIMIT:
        raise BusinessException(ErrorCode.INVENTORY_XLSX_LIMIT_EXCEEDED)

    rows = await inventory_crud.list_inventory_items_by_codes(db, codes)
    row_by_code = {item.inventory_code: (item, grade) for item, grade in rows}
    if len(row_by_code) != len(codes):
        raise BusinessException(ErrorCode.INVENTORY_ITEM_NOT_FOUND)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory"
    worksheet.append([header for _, header in INVENTORY_EXPORT_HEADERS])
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 18

    for row_index, code in enumerate(codes, start=2):
        item, grade = row_by_code[code]
        worksheet.cell(row=row_index, column=1, value=item.inventory_code)
        worksheet.cell(row=row_index, column=2, value="")
        worksheet.cell(row=row_index, column=3, value=item.width)
        worksheet.cell(row=row_index, column=4, value=item.length)
        worksheet.cell(row=row_index, column=5, value=grade)
        worksheet.cell(row=row_index, column=6, value=item.thickness)
        worksheet.cell(row=row_index, column=7, value=item.quantity)
        worksheet.cell(row=row_index, column=8, value="")
        worksheet.cell(row=row_index, column=9, value="")

    output = BytesIO()
    workbook.save(output)
    logger.info("inventory_xlsx_exported row_count={}", len(codes))
    return output.getvalue()
