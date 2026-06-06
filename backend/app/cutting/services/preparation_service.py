from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.cutting.cruds import preparation_crud
from app.cutting.models.preparation import (
    CuttingPreparationItem,
    CuttingPreparationOrder,
    CuttingTemplateExport,
)
from app.cutting.schemas.preparation import (
    PreparationItemCreateIn,
    PreparationItemOut,
    PreparationOrderCreateIn,
    PreparationOrderOut,
    TemplateExportOut,
)
from app.material.cruds import inventory_crud, material_crud
from common.error_codes import ErrorCode
from common.exceptions import BusinessException
from settings import get_project_root, get_settings, get_site_runtime_root

TEMPLATE_HEADERS = ["板材名称", "图纸路径", "宽", "长", "材质", "厚度", "数量"]
BACKEND_ROOT = get_site_runtime_root()
PROJECT_ROOT = get_project_root()
TEMPLATE_SAMPLE = PROJECT_ROOT / "resources" / "Template.xlsx"


def _to_item_out(item: CuttingPreparationItem) -> PreparationItemOut:
    return PreparationItemOut.model_validate(item)


def _to_export_out(export: CuttingTemplateExport) -> TemplateExportOut:
    return TemplateExportOut(
        id=export.id,
        order_id=export.order_id,
        file_name=export.file_name,
        file_path=export.file_path,
        row_count=export.row_count,
        created_by=export.created_by,
        created_at=export.created_at,
        download_url=f"/api/v1/cutting-preparations/template-exports/{export.id}/download",
    )


async def _to_order_out(
    db: AsyncSession,
    order: CuttingPreparationOrder,
) -> PreparationOrderOut:
    items = await preparation_crud.list_items(db, order.id)
    return PreparationOrderOut(
        id=order.id,
        preparation_date=order.preparation_date,
        status=order.status,
        created_by=order.created_by,
        exported_file_id=order.exported_file_id,
        items=[_to_item_out(item) for item in items],
    )


async def create_order(
    db: AsyncSession,
    data: PreparationOrderCreateIn,
    *,
    actor: str,
) -> PreparationOrderOut:
    order = await preparation_crud.create_order(
        db,
        CuttingPreparationOrder(
            preparation_date=data.preparation_date,
            status="draft",
            created_by=actor,
        ),
    )
    await db.commit()
    return await _to_order_out(db, order)


async def list_orders(db: AsyncSession) -> list[PreparationOrderOut]:
    orders = await preparation_crud.list_orders(db)
    return [await _to_order_out(db, order) for order in orders]


async def get_order(db: AsyncSession, order_id: int) -> PreparationOrderOut:
    order = await preparation_crud.get_order(db, order_id)
    if order is None:
        raise BusinessException(ErrorCode.PREPARATION_ORDER_NOT_FOUND)
    return await _to_order_out(db, order)


async def _validate_source_inventory(
    db: AsyncSession,
    data: PreparationItemCreateIn,
) -> None:
    if data.source_inventory_item_id is None:
        return

    item = await inventory_crud.get_inventory_item(db, data.source_inventory_item_id)
    if item is None or item.status != "available" or not item.reusable:
        raise BusinessException(ErrorCode.INVALID_PREPARATION_SOURCE)
    material = await material_crud.get_material(db, item.material_id)
    if material is None:
        raise BusinessException(ErrorCode.INVALID_PREPARATION_SOURCE)
    if material.material_grade != data.material_grade:
        raise BusinessException(ErrorCode.INVALID_PREPARATION_SOURCE)
    if item.thickness != data.thickness or item.width < data.width or item.length < data.length:
        raise BusinessException(ErrorCode.INVALID_PREPARATION_SOURCE)
    if item.quantity < data.quantity:
        raise BusinessException(ErrorCode.INVALID_PREPARATION_SOURCE)
    item.status = "reserved"


async def add_item(
    db: AsyncSession,
    order_id: int,
    data: PreparationItemCreateIn,
) -> PreparationOrderOut:
    order = await preparation_crud.get_order(db, order_id)
    if order is None:
        raise BusinessException(ErrorCode.PREPARATION_ORDER_NOT_FOUND)
    if order.status != "draft":
        raise BusinessException(ErrorCode.INVALID_PREPARATION_STATUS)

    await _validate_source_inventory(db, data)
    await preparation_crud.create_item(
        db,
        CuttingPreparationItem(order_id=order_id, **data.model_dump()),
    )
    await db.commit()
    return await _to_order_out(db, order)


def _get_export_dir() -> Path:
    storage_dir = Path(get_settings().storage_dir)
    if not storage_dir.is_absolute():
        storage_dir = BACKEND_ROOT / storage_dir
    export_dir = storage_dir / "exports" / "templates"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _build_workbook() -> Workbook:
    if TEMPLATE_SAMPLE.exists():
        workbook = load_workbook(TEMPLATE_SAMPLE)
        sheet = workbook.active
        for column, header in enumerate(TEMPLATE_HEADERS, start=1):
            sheet.cell(row=1, column=column, value=header)
        return workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(TEMPLATE_HEADERS)
    return workbook


async def export_template(
    db: AsyncSession,
    order_id: int,
    *,
    actor: str,
) -> TemplateExportOut:
    order = await preparation_crud.get_order(db, order_id)
    if order is None:
        raise BusinessException(ErrorCode.PREPARATION_ORDER_NOT_FOUND)
    items = await preparation_crud.list_items(db, order_id)

    workbook = _build_workbook()
    sheet = workbook.active
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for item in items:
        sheet.append(
            [
                item.sheet_name,
                item.drawing_path,
                item.width,
                item.length,
                item.material_grade,
                item.thickness,
                item.quantity,
            ]
        )

    file_name = f"cutting-template-order-{order_id}.xlsx"
    file_path = _get_export_dir() / file_name
    workbook.save(file_path)

    export = await preparation_crud.create_export(
        db,
        CuttingTemplateExport(
            order_id=order_id,
            file_name=file_name,
            file_path=str(file_path),
            row_count=len(items),
            created_by=actor,
        ),
    )
    order.status = "generated"
    order.exported_file_id = export.id
    await db.commit()
    await db.refresh(export)
    return _to_export_out(export)


async def get_export_file(
    db: AsyncSession,
    export_id: int,
) -> tuple[Path, str]:
    export = await preparation_crud.get_export(db, export_id)
    if export is None:
        raise BusinessException(ErrorCode.TEMPLATE_EXPORT_NOT_FOUND)
    file_path = Path(export.file_path)
    if not file_path.exists():
        raise BusinessException(ErrorCode.TEMPLATE_EXPORT_NOT_FOUND)
    return file_path, export.file_name
